

import pandas as pd
from openpyxl import load_workbook
#read the excel if it is present

path = 'Lost_and_Found.xlsx'
try:
    
    with open(path, 'rb') as f:
        df = pd.read_excel(f, 0)
    
    
#create a excel file
except:    
    df = pd.DataFrame(columns= 
                      ['Item', 'Date', 'Color', 'Location', 'Note', 'Check'])
    df.to_excel(path, index= False)
    wb= load_workbook(path)
    ws= wb['Sheet1']
    ws.column_dimensions['A'].width= 10
    ws.column_dimensions['B'].width= 10
    ws.column_dimensions['C'].width= 8
    ws.column_dimensions['D'].width= 15
    ws.column_dimensions['E'].width= 30
    ws.column_dimensions['F'].width= 8
    wb.save(path)
    wb.close()
   

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

def write(item:str, date:str, color:str, location:str, note:str):
    df =pd.DataFrame({'Item':[item], 'Date':[date], 'Color':[color], 
                      'Location':[location], 'Note':[note]})
    wb = load_workbook(filename=path)
    ws = wb['Sheet1'] 
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    wb.save(path)
    wb.close()

    
    
def find(item:str, color:str):
    with open(path, 'rb') as f:
        df = pd.read_excel(f)
        res= df.loc[(df['Item'] == item) & (df['Color'] == color) & 
                (pd.isna(df['Check'])), ['Item', 'Date', 'Color', 'Location', 'Note']]
        return res    
    

from tkinter import *
from tkinter import ttk
from tkcalendar import DateEntry

window = Tk()
window.title("Lost and Found System")
window.geometry('500x250+400+200')
itemvalue= ('Bottle', 'Umbrella', 'Coat','laptop','Cellphone', 'Wallet',
            'Charger', 'Bag', 'Mug or Cup', 'Others') 
colorvalue= ('Pink','Red','Orange','Yellow','Green','Blue','Dark Blue','Purple',
             'Brown', 'Gray', 'Black','White','Beige')
locationvalue= ('First Floor', 'Second Floor', 'Thrid Floor')


def searchingWindow():
    #the window shows the result
    def searchvalue():
        item = itemV.get()
        color = colorV.get()
        
        reswindow = Toplevel(window)
        reswindow.title('Lost and Found System')
        reswindow.geometry('400x525+430+90')
        
        if find(item, color).empty:
            Label(reswindow, text='No matching item', font=('Times New Roman'
                ,14)).place(relx=0.5, rely=0.2, anchor='center')
        elif find(item, color).any:
            Label(reswindow, text=str(find(item, color)), justify= 'center', width=40,
                  font=('Times New Roman',9)).pack()
        Button(reswindow, text='Back to Menu', command= lambda: [window.deiconify(), 
                reswindow.destroy()]).place(relx=0.39, rely=0.95, anchor='center')
        Button(reswindow, text='        Exit        ', command= window.destroy).place(
            relx=0.61, rely=0.95, anchor='center')
        
    #the window asks for characteristics to search
    searchwindow = Toplevel(window)
    searchwindow.title('Lost and Found System')
    searchwindow.geometry('500x250+400+200')
    
    Label(searchwindow, text='Select the Characteristics',
          font=('Times New Roman', 18)).place(relx=0.5,rely=0.13, anchor='center')
    
    itemV = StringVar()
    Label(searchwindow, text='Item',font=('Times New Roman', 14)).place(
        relx=0.2, rely=0.38, anchor='w')
    itemS= ttk.Combobox(searchwindow,justify='center', textvariable= itemV,
                        width= 10,font=('Times New Roman', 12))
    itemS['values']= itemvalue
    itemS.place(relx=0.68, rely=0.38, anchor='center')
    itemS.current(0)
    
    colorV = StringVar()
    Label(searchwindow, text='Color',font=('Times New Roman', 14)).place(
        relx=0.2, rely= 0.53, anchor='w')
    colorS= ttk.Combobox(searchwindow,justify='center', textvariable= colorV,
                         width= 10,font=('Times New Roman', 12))
    colorS['values']= colorvalue
    colorS.place(relx=0.68, rely=0.53, anchor='center')
    colorS.current(0)
    
    Button(searchwindow, text='     Confirm     ', font=('Times New Roman', 11), 
        command= lambda: [searchvalue(),searchwindow.destroy()]).place(relx=0.6, 
        rely=0.78, anchor='center')
    Button(searchwindow, text='Back to Menu', font=('Times New Roman', 11), 
        command= lambda: [searchwindow.destroy(),window.deiconify()]).place(
        relx=0.4, rely=0.78,anchor='center')
    Button(searchwindow, text='      Exit      ', font=('Times New Roman', 11), 
           command= window.destroy).place(relx=0.5,rely=0.91, anchor='center')


    
    # the window reminds successful submission
def submittedWindow():
    # Toplevel object which will
    # be treated as a new window
    submitWindow = Toplevel(window)
 
    # sets the title of the
    # Toplevel widget
    submitWindow.title("Lost and Found System")
 
    # sets the geometry of toplevel
    submitWindow.geometry("400x250+430+200")
 
    Label(submitWindow, text ="Successfully Submitted",font=('Times New Roman', 
        18)).place(relx=0.5, rely=0.3, anchor='center')
    Button(submitWindow, text='Back to Menu', font=('Times New Roman', 12),
        command= lambda:[window.deiconify(), submitWindow.destroy()]).place(
            relx=0.35,rely=0.7, anchor='center')
    Button(submitWindow, text='        Exit        ', font=('Times New Roman', 12),
        command= window.destroy).place(relx= 0.65, rely=0.7, anchor='center')


def inputWindow():
    # get value from combobox
    def getvalue():
        item = itemV.get()
        date = dateV.get()
        color = colorV.get()
        location = locationV.get()
        note = noteV.get()

        write(item, date, color, location, str(note))
                    
        
    datawindow = Toplevel(window)
    datawindow.title('Lost and Found System')
    datawindow.geometry('450x350+430+150')    
    Label(datawindow, text='Input the Characteristics',font=('Times New Roman', 18)
          ).place(relx=0.5, rely=0.08, anchor='center')
    #items
    itemV = StringVar()
    Label(datawindow, text= 'Item', justify= RIGHT, 
          font=('Times New Roman', 14)).place(relx=0.15, rely= 0.25, anchor='w')
    
    itemIn= ttk.Combobox(datawindow,justify='center', textvariable= itemV, width= 15,font=('Times New Roman', 12))
    itemIn['values']= itemvalue
    itemIn.place(relx=0.73,rely=0.25, anchor='center')
    itemIn.current(0)
    
    
    #date
    dateV = StringVar()
    Label(datawindow,text='Date',justify= RIGHT,font=('Times New Roman', 
              14)).place(relx= 0.15, rely=0.35, anchor='w')
    DateEntry(datawindow,justify='center', selectmode='day', textvariable= dateV, 
              width= 15,font=('Times New Roman', 12)).place(relx= 0.73, rely= 0.35, anchor='center')
    
    #color
    colorV = StringVar()
    Label(datawindow, text= 'Color', justify= RIGHT,font=('Times New Roman', 
                14)).place(relx=0.15, rely=0.45, anchor='w')
    colorIn= ttk.Combobox(datawindow,justify='center', textvariable= colorV, width= 15,font=('Times New Roman', 12))
    colorIn['values'] = colorvalue
    colorIn.place(relx=0.73, rely=0.45, anchor='center')
    colorIn.current(0)
    
    #location
    locationV = StringVar() 
    Label(datawindow, text= 'Location', justify= RIGHT,font=('Times New Roman', 
            14)).place(relx=0.15, rely=0.55, anchor='w')
    locaIn= ttk.Combobox(datawindow, justify='center', textvariable= locationV, width= 15,font=('Times New Roman', 12))
    locaIn['values']= locationvalue
    locaIn.place(relx=0.73, rely=0.55, anchor='center')
    locaIn.current(0)
    
    #note
    noteV = StringVar()
    Label(datawindow, text= 'Note', justify= RIGHT, font=('Times New Roman', 
            14)).place(relx=0.15, rely=0.65, anchor='w')
    ttk.Entry(datawindow, textvariable= noteV, width= 20).place(relx=0.73, 
            rely=0.65, anchor='center')
    
    
    Button(datawindow, text='       Confirm     ', font=('Times New Roman', 12), 
        command= lambda: [getvalue(),submittedWindow(),datawindow.destroy()]
        ).place(relx=0.6, rely=0.81, anchor='center')
    Button(datawindow, text='Back to menu', font=('Times New Roman', 12), 
        command= lambda: [datawindow.destroy(),window.deiconify()]
        ).place(relx=0.40, rely= 0.81, anchor='center')
    Button(datawindow, text='   Exit   ', font=('Times New Roman', 12),
        command= lambda:[datawindow.destroy(),window.destroy()]
        ).place(relx=0.5, rely=0.91, anchor='center')


Label(window, text='Wlecome to Lost and Found System',font=('Times New Roman', 18)
      ).place(relx=0.5, rely=0.2, anchor='center')
Button(window, text='Submit an Item', font=('Times New Roman', 13), 
       justify= CENTER, command= lambda:[inputWindow(),window.withdraw()]
       ).place(relx=0.38, rely= 0.65, anchor='center')
Button(window, text='   Find an Item   ', font=('Times New Roman', 13), 
       justify= CENTER, command= lambda:[searchingWindow(), window.withdraw()]
       ).place(relx=0.62, rely= 0.65, anchor='center')
Button(window, text= '    Exit    ', font=('Times New Roman', 12), 
       command= window.destroy).place(relx=0.5, rely=0.8, anchor='center')


window.mainloop()
